import streamlit as st
import openpyxl
import camelot
import fitz  # PyMuPDF
import re
import os
import json
from io import BytesIO
from openpyxl import load_workbook

def should_remove_column(column_data, ignore_indices):
    threshold = 0.9
    # Use enumerate to keep track of the index
    non_empty_cells = [(index, cell) for index, cell in enumerate(column_data) if cell.strip() != '']
    # Adjust logic to check if the index is in ignore_indices
    checkmark_or_na_cells = [(index, cell) for index, cell in non_empty_cells if (cell == '✔' or cell == 'N/A') and (index not in ignore_indices)]

    # Ensure there are non-empty cells before dividing to prevent division by zero
    if len(non_empty_cells) == 0:
        return False
    return all(cell.strip() == '' for cell in column_data[0:2]) 

def process_pdf_table(pdf_path, page_number, ignore_na_indices):
    # Coordinates for column divisions, as a comma-separated string
    columns_coords = "0,60,151,210,284,353,413,450,550,585"
    # Table areas for the first page and subsequent pages
    table_area_first_page = "0,680,310,550"
    table_area_other_pages = "0,700,600,50"

    # Configure for the first page, using 'table_areas'
    if page_number == 1:
        tables = camelot.read_pdf(pdf_path, pages="1", flavor='stream', table_areas=[table_area_first_page])
    else:
        # For other pages, use both 'table_areas' and 'columns'
        tables = camelot.read_pdf(pdf_path, pages=str(page_number), flavor='stream', table_areas=[table_area_other_pages], columns=[columns_coords])
    
    # Verify table extraction
    if tables.n == 0:
        print(f"No tables found on page {page_number}.")
        return []

    # Process the extracted data
    processed_data = []
    for table in tables:
        data = [list(row) for row in table.df.itertuples(index=False)]
        # This transposes the table data to work with columns
        columns = list(zip(*data))
        # Determine which columns to remove
        columns_to_remove = [index for index, col in enumerate(columns) if index not in ignore_na_indices and should_remove_column(col, ignore_na_indices)]
        # Remove the marked columns
        for index in sorted(columns_to_remove, reverse=True):
            for row in data:
                del row[index]
        processed_data.extend(data)

    return processed_data

def extract_text_for_filename(pdf_path):
    doc = fitz.open(pdf_path)
    page = doc[0]
    rect = fitz.Rect(438.10, 62.91, 603.96, 88.87)
    text = page.get_textbox(rect)
    match = re.search(r'C\d+', text)
    cert_number = match.group(0) if match else "extracted_data"
    doc.close()
    return cert_number

def save_data_to_excel(data_per_page):
    workbook = openpyxl.Workbook()
    del workbook['Sheet']
    
    for index, page_data in enumerate(data_per_page, start=1):
        sheet = workbook.create_sheet(title=f"Page {index}")
        for row in page_data:
            sheet.append(row)
    
    # En lugar de guardar el archivo, lo devolvemos como un objeto BytesIO
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output



#XLSX TO JSON


def process_first_page(sheet):
    data = {}
    for row in sheet.iter_rows(min_row=1, max_col=2, values_only=True):
        key = row[0].strip() if row[0] else None
        value = row[1].strip() if row[1] else None
        if key and value:
            data[key] = value
    return data

def process_subsequent_pages(sheet):
    data = {}
    current_group = None
    current_group_data = []
    column_names = []
    group_start = False

    exclude_lines = [
        "After Adjustment", "End of datasheet",
        "Compliance/Pass", "Non-compliance/Fail",
        "Accuracy Chart",  # Clave añadida como solicitado.
        # Puedes agregar cualquier otra línea o palabras clave que desees excluir aquí.
    ]

    for row in sheet.iter_rows(values_only=True):
        row_values = [cell if cell is not None else "" for cell in row]
        cleaned_row_values = []

        # Limpiar los valores de las filas y dividir las celdas combinadas si es necesario.
        for cell in row_values:
            cell_value = str(cell).strip()
            if "Units Max. Error(Tol.)" in cell_value:
                # Dividir en "Units" y "Max. Error(Tol.)"
                parts = re.split(r'\s(?=Max. Error\(Tol.\))', cell_value)
                cleaned_row_values.extend(parts)
            else:
                cleaned_row_values.append(cell_value)

        # Eliminar valores nulos o líneas excluidas
        row_values = [val for val in cleaned_row_values if val and all(exclude not in val for exclude in exclude_lines)]

        # Si la fila está vacía después de filtrar, continuar al siguiente ciclo
        if not row_values:
            continue

        # Detectar inicio de nuevo grupo
        if len(row_values) == 1 and row_values[0] not in exclude_lines:
            if current_group and current_group_data:
                # Si ya hay un grupo y datos capturados, guardar antes de empezar uno nuevo
                data[current_group] = current_group_data
                current_group_data = []
            current_group = row_values[0].lower().replace(' ', '_')
            group_start = True
            continue

        # Si se detecta inicio de grupo, establecer nombres de columna
        if group_start:
            column_names = [x.lower().replace(' ', '_') for x in row_values]
            group_start = False
            continue

        # Recolectar datos de las mediciones
        if current_group and column_names:
            # Asegurarse de que la fila actual no es un inicio de un nuevo grupo o una línea excluida
            if not any(exclude in row_values[0] for exclude in exclude_lines) and not group_start:
                row_data = {}
                for i, column_name in enumerate(column_names):
                    if i < len(row_values):
                        row_data[column_name] = row_values[i]
                    else:
                        # Si no hay suficientes valores, usar "N/A"
                        row_data[column_name] = "N/A"
                current_group_data.append(row_data)

    # Añadir el último grupo recogido fuera del bucle, si hay alguno
    if current_group and current_group_data:
        data[current_group] = current_group_data

    return data

def process_workbook_from_stream(excel_stream):
    # Cargar el workbook directamente desde un stream de BytesIO
    workbook = load_workbook(excel_stream)
    first_page_data = process_first_page(workbook.worksheets[0])
    subsequent_pages_data = {}
    for sheet in workbook.worksheets[1:]:
        sheet_data = process_subsequent_pages(sheet)
        if sheet_data:
            subsequent_pages_data.update(sheet_data)
    return {
        "datasheet_info": first_page_data,
        "measurements": subsequent_pages_data
    }

def update_certificate_data(cert_number, workbook_data):
    certificate_data_file = 'certificate_data.json'
    
    # Intentar leer los datos existentes, o inicializar a vacío si el archivo no existe
    if os.path.isfile(certificate_data_file):
        with open(certificate_data_file, 'r', encoding="utf-8") as file:
            certificate_data = json.load(file)
    else:
        certificate_data = {}
    
    # Preparar los datos específicos del certificado para añadir/actualizar
    certificate_specific_data = {"CertNo": cert_number}
    certificate_specific_data.update(workbook_data)
    
    # Actualizar o añadir los datos del nuevo certificado
    certificate_data[cert_number] = certificate_specific_data
    
    # Guardar los datos actualizados en el archivo
    with open(certificate_data_file, 'w', encoding="utf-8") as file:
        json.dump(certificate_data, file, ensure_ascii=False, sort_keys=True, indent=2)



def main():
    st.title("PDF Table Certification Extractor")

    uploaded_pdf = st.file_uploader("Upload your PDF here", type='pdf')

    if uploaded_pdf is not None:
        # Crear un archivo temporal para el PDF subido
        temp_pdf_path = "temp_uploaded_file.pdf"
        with open(temp_pdf_path, "wb") as f:
            f.write(uploaded_pdf.getvalue())

        # Extraer el número de certificado del PDF
        cert_number = extract_text_for_filename(temp_pdf_path)

        ignore_na_columns = [1, 2]
        all_processed_data = []
        doc = fitz.open(temp_pdf_path)
        
        # Procesar cada página del PDF para extracción de tablas
        for page_num in range(1, len(doc) + 1):
            page_data = process_pdf_table(temp_pdf_path, page_num, ignore_na_columns)
            all_processed_data.append(page_data)
        doc.close()
        
        # Guardar los datos procesados en un archivo XLSX en memoria
        excel_stream = save_data_to_excel(all_processed_data)
        
        try:
            os.remove(temp_pdf_path)  # Eliminar el archivo PDF temporal
        except PermissionError as e:
            st.error(f"Error al eliminar el archivo temporal: {e}")

        # Ofrecer la descarga del archivo XLSX
        st.download_button(
            label="Download Excel",
            data=excel_stream,
            file_name=f"{cert_number}_extracted_tables.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        # Reposicionar el puntero del stream de Excel para su lectura
        excel_stream.seek(0)
        
        # Procesar el archivo XLSX en memoria para convertirlo a JSON
        workbook_data = process_workbook_from_stream(excel_stream)
        
        # Generar el objeto BytesIO para el archivo JSON
        json_stream = BytesIO()
        json_stream.write(json.dumps(workbook_data, ensure_ascii=False, indent=2).encode('utf-8'))
        json_stream.seek(0)

        # Ofrecer la descarga del archivo JSON
        st.download_button(
            label="Download JSON",
            data=json_stream,
            file_name=f"{cert_number}_extracted_data.json",
            mime="application/json"
        )

        # Actualizar el archivo acumulativo certificate_data.json con los nuevos datos
        update_certificate_data(cert_number, workbook_data)

if __name__ == "__main__":
    main()